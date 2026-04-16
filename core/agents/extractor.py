"""Visual Metadata Extractor Agent.

This agent scans Excel files and extracts structural metadata including:
- Cell colors and their semantic meanings
- Merged cell ranges and their values
- Section boundaries based on visual cues
- Empty row/column detection for section breaks
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from core.agents.base import AgentResult, BaseAgent
from core.logging import get_logger

logger = get_logger(__name__)


# Common color mappings for semantic meaning
COLOR_LABELS = {
    "FF0000": "Red/Alert",
    "00FF00": "Green/Success",
    "0000FF": "Blue/Info",
    "FFFF00": "Yellow/Warning",
    "FFA500": "Orange/Caution",
    "800080": "Purple/Special",
    "008000": "DarkGreen/Verified",
    "000080": "Navy/Header",
    "808080": "Gray/Disabled",
    "FFFFFF": "White/Default",
    "000000": "Black/Text",
    "D3D3D3": "LightGray/Separator",
    "ADD8E6": "LightBlue/Highlight",
    "90EE90": "LightGreen/Positive",
    "FFB6C1": "LightPink/Attention",
    "FFFACD": "LemonChiffon/Note",
}


@dataclass
class CellMetadata:
    """Metadata for a single cell."""

    row: int
    column: int
    column_letter: str
    value: Any
    data_type: str
    background_color: str | None
    color_label: str | None
    font_color: str | None
    font_bold: bool
    font_size: float | None
    is_merged: bool
    merged_range: str | None
    formula: str | None
    number_format: str | None


@dataclass
class MergedRegion:
    """Information about a merged cell region."""

    range_string: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    value: Any
    background_color: str | None
    color_label: str | None
    row_span: int
    col_span: int


@dataclass
class SheetSection:
    """A logical section within a sheet identified by visual cues."""

    name: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    header_color: str | None
    color_label: str | None
    column_headers: list[str]
    row_count: int
    has_data: bool


@dataclass
class SheetManifest:
    """Complete structural manifest for a single sheet."""

    sheet_name: str
    total_rows: int
    total_cols: int
    merged_regions: list[MergedRegion] = field(default_factory=list)
    sections: list[SheetSection] = field(default_factory=list)
    color_regions: dict[str, list[tuple[int, int]]] = field(default_factory=dict)
    empty_rows: list[int] = field(default_factory=list)
    empty_cols: list[int] = field(default_factory=list)
    header_row: int | None = None
    data_start_row: int | None = None
    sample_data: list[dict] = field(default_factory=list)


@dataclass
class WorkbookManifest:
    """Complete structural manifest for an entire workbook."""

    file_path: str
    sheet_count: int
    sheet_names: list[str]
    sheets: dict[str, SheetManifest] = field(default_factory=dict)
    total_merged_regions: int = 0
    total_sections: int = 0
    detected_colors: list[str] = field(default_factory=list)


class VisualMetadataExtractor(BaseAgent):
    """Agent that extracts visual metadata from Excel files.

    This agent scans Excel files using openpyxl to extract:
    - Cell colors and their semantic meanings
    - Merged cell ranges with broadcasted values
    - Section boundaries based on visual patterns
    - Empty row/column detection for structure analysis
    """

    def __init__(self):
        """Initialize the Visual Metadata Extractor agent."""
        super().__init__(name="VisualMetadataExtractor")
        self._empty_threshold = 0.9  # 90% empty = consider row/col empty

    async def execute(
        self,
        file_path: str | None = None,
        file_content: bytes | None = None,
        include_sample_data: bool = True,
        sample_rows: int = 5,
    ) -> AgentResult:
        """Extract visual metadata from an Excel file.

        Args:
            file_path: Path to the Excel file.
            file_content: Raw bytes of the Excel file (alternative to path).
            include_sample_data: Whether to include sample data rows.
            sample_rows: Number of sample rows to include.

        Returns:
            AgentResult containing WorkbookManifest or error.

        """
        self._log_start({"file_path": file_path})

        try:
            if file_content:
                workbook = openpyxl.load_workbook(
                    filename=BytesIO(file_content),
                    data_only=False,  # Keep formulas for analysis
                )
            elif file_path:
                workbook = openpyxl.load_workbook(
                    filename=file_path,
                    data_only=False,
                )
            else:
                return AgentResult(
                    success=False,
                    error="Either file_path or file_content must be provided",
                )

            manifest = self._extract_workbook_manifest(
                workbook=workbook,
                file_path=file_path or "in-memory",
                include_sample_data=include_sample_data,
                sample_rows=sample_rows,
            )

            workbook.close()

            result = AgentResult(
                success=True,
                data=manifest,
                metadata={
                    "sheet_count": manifest.sheet_count,
                    "total_merged_regions": manifest.total_merged_regions,
                    "total_sections": manifest.total_sections,
                },
            )
            self._log_complete(result)
            return result

        except Exception as e:
            self._log_error(e)
            return AgentResult(
                success=False,
                error=f"Failed to extract metadata: {str(e)}",
            )

    def _extract_workbook_manifest(
        self,
        workbook: openpyxl.Workbook,
        file_path: str,
        include_sample_data: bool,
        sample_rows: int,
    ) -> WorkbookManifest:
        """Extract manifest for entire workbook."""
        manifest = WorkbookManifest(
            file_path=file_path,
            sheet_count=len(workbook.sheetnames),
            sheet_names=list(workbook.sheetnames),
        )

        all_colors = set()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_manifest = self._extract_sheet_manifest(
                sheet=sheet,
                include_sample_data=include_sample_data,
                sample_rows=sample_rows,
            )
            manifest.sheets[sheet_name] = sheet_manifest
            manifest.total_merged_regions += len(sheet_manifest.merged_regions)
            manifest.total_sections += len(sheet_manifest.sections)

            # Collect all unique colors
            all_colors.update(sheet_manifest.color_regions.keys())

        manifest.detected_colors = list(all_colors)
        return manifest

    def _extract_sheet_manifest(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        include_sample_data: bool,
        sample_rows: int,
    ) -> SheetManifest:
        """Extract manifest for a single sheet."""
        manifest = SheetManifest(
            sheet_name=sheet.title,
            total_rows=sheet.max_row or 0,
            total_cols=sheet.max_column or 0,
        )

        if manifest.total_rows == 0 or manifest.total_cols == 0:
            return manifest

        # Extract merged regions
        manifest.merged_regions = self._extract_merged_regions(sheet)

        # Extract color regions
        manifest.color_regions = self._extract_color_regions(sheet)

        # Detect empty rows and columns
        manifest.empty_rows = self._detect_empty_rows(sheet)
        manifest.empty_cols = self._detect_empty_cols(sheet)

        # Detect sections based on visual patterns
        manifest.sections = self._detect_sections(sheet, manifest)

        # Find header row and data start
        header_info = self._find_header_row(sheet)
        manifest.header_row = header_info.get("row")
        manifest.data_start_row = header_info.get("data_start")

        # Include sample data if requested
        if include_sample_data and manifest.data_start_row:
            manifest.sample_data = self._extract_sample_data(
                sheet=sheet,
                start_row=manifest.data_start_row,
                header_row=manifest.header_row,
                num_rows=sample_rows,
            )

        return manifest

    def _extract_merged_regions(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[MergedRegion]:
        """Extract all merged cell regions with their values."""
        merged_regions = []

        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            cell = sheet.cell(row=min_row, column=min_col)

            bg_color = self._get_cell_background_color(cell)
            color_label = self._color_to_label(bg_color) if bg_color else None

            region = MergedRegion(
                range_string=str(merged_range),
                start_row=min_row,
                end_row=max_row,
                start_col=min_col,
                end_col=max_col,
                value=cell.value,
                background_color=bg_color,
                color_label=color_label,
                row_span=max_row - min_row + 1,
                col_span=max_col - min_col + 1,
            )
            merged_regions.append(region)

        return merged_regions

    def _extract_color_regions(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[tuple[int, int]]]:
        """Extract cells grouped by background color."""
        color_regions: dict[str, list[tuple[int, int]]] = {}

        for row_idx in range(1, min(sheet.max_row + 1, 1001)):  # Limit to 1000 rows
            for col_idx in range(1, min(sheet.max_column + 1, 51)):  # Limit to 50 cols
                cell = sheet.cell(row=row_idx, column=col_idx)
                bg_color = self._get_cell_background_color(cell)

                if bg_color and bg_color not in ("FFFFFF", "000000", None):
                    color_label = self._color_to_label(bg_color)
                    if color_label not in color_regions:
                        color_regions[color_label] = []
                    color_regions[color_label].append((row_idx, col_idx))

        return color_regions

    def _detect_empty_rows(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[int]:
        """Detect rows that are mostly empty (section breaks)."""
        empty_rows = []

        for row_idx in range(1, sheet.max_row + 1):
            empty_count = 0
            total_cols = sheet.max_column

            for col_idx in range(1, total_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "":
                    empty_count += 1

            if total_cols > 0 and (empty_count / total_cols) >= self._empty_threshold:
                empty_rows.append(row_idx)

        return empty_rows

    def _detect_empty_cols(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[int]:
        """Detect columns that are mostly empty."""
        empty_cols = []

        for col_idx in range(1, sheet.max_column + 1):
            empty_count = 0
            total_rows = sheet.max_row

            for row_idx in range(1, total_rows + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None or str(cell.value).strip() == "":
                    empty_count += 1

            if total_rows > 0 and (empty_count / total_rows) >= self._empty_threshold:
                empty_cols.append(col_idx)

        return empty_cols

    def _detect_sections(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        manifest: SheetManifest,
    ) -> list[SheetSection]:
        """Detect logical sections based on merged cells and empty rows."""
        sections = []

        # Use merged regions that span multiple columns as section headers
        header_regions = [mr for mr in manifest.merged_regions if mr.col_span >= 3 and mr.value is not None]

        # Sort by row
        header_regions.sort(key=lambda x: x.start_row)

        for i, header in enumerate(header_regions):
            # Determine section end (next header or end of data)
            if i + 1 < len(header_regions):
                end_row = header_regions[i + 1].start_row - 1
            else:
                end_row = sheet.max_row

            # Find actual data range within section
            data_start = header.end_row + 1
            col_headers = self._get_column_headers(sheet, data_start, header.start_col, header.end_col)

            section = SheetSection(
                name=str(header.value) if header.value else f"Section_{i + 1}",
                start_row=header.start_row,
                end_row=end_row,
                start_col=header.start_col,
                end_col=header.end_col,
                header_color=header.background_color,
                color_label=header.color_label,
                column_headers=col_headers,
                row_count=end_row - data_start + 1 if data_start <= end_row else 0,
                has_data=data_start <= end_row,
            )
            sections.append(section)

        # If no sections detected from merged cells, try to detect from color patterns
        if not sections and manifest.color_regions:
            sections = self._detect_sections_from_colors(sheet, manifest)

        return sections

    def _detect_sections_from_colors(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        manifest: SheetManifest,
    ) -> list[SheetSection]:
        """Detect sections based on colored header rows."""
        sections = []
        processed_rows = set()

        for color_label, cells in manifest.color_regions.items():
            # Find rows where many cells have the same color (likely headers)
            row_counts: dict[int, int] = {}
            for row, col in cells:
                row_counts[row] = row_counts.get(row, 0) + 1

            # Rows with 3+ colored cells might be section headers
            for row, count in row_counts.items():
                if count >= 3 and row not in processed_rows:
                    processed_rows.add(row)

                    # Get the header text
                    header_text = None
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col_idx)
                        if cell.value:
                            header_text = str(cell.value)
                            break

                    section = SheetSection(
                        name=header_text or f"Section_Row_{row}",
                        start_row=row,
                        end_row=row,  # Will be updated later
                        start_col=1,
                        end_col=sheet.max_column,
                        header_color=color_label,
                        color_label=color_label,
                        column_headers=[],
                        row_count=0,
                        has_data=False,
                    )
                    sections.append(section)

        # Sort sections by row and calculate end rows
        sections.sort(key=lambda x: x.start_row)
        for i, section in enumerate(sections):
            if i + 1 < len(sections):
                section.end_row = sections[i + 1].start_row - 1
            else:
                section.end_row = sheet.max_row

            section.row_count = section.end_row - section.start_row
            section.has_data = section.row_count > 0

        return sections

    def _find_header_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int | None]:
        """Find the header row by looking for rows with many text values."""
        result: dict[str, int | None] = {"row": None, "data_start": None}

        for row_idx in range(1, min(sheet.max_row + 1, 20)):  # Check first 20 rows
            text_count = 0
            non_empty_count = 0

            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    non_empty_count += 1
                    if isinstance(cell.value, str) and not cell.value.replace(".", "").isdigit():
                        text_count += 1

            # Header row typically has many text values
            if non_empty_count >= 3 and text_count >= non_empty_count * 0.6:
                result["row"] = row_idx
                result["data_start"] = row_idx + 1
                break

        return result

    def _get_column_headers(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        start_col: int,
        end_col: int,
    ) -> list[str]:
        """Get column headers from a specific row."""
        headers = []
        for col_idx in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col_idx)
            headers.append(str(cell.value) if cell.value else f"Column_{get_column_letter(col_idx)}")
        return headers

    def _extract_sample_data(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        start_row: int,
        header_row: int | None,
        num_rows: int,
    ) -> list[dict]:
        """Extract sample data rows for preview."""
        sample_data = []

        # Get headers
        if header_row:
            headers = [
                str(sheet.cell(row=header_row, column=col).value) or f"Col_{col}" for col in range(1, sheet.max_column + 1)
            ]
        else:
            headers = [f"Col_{col}" for col in range(1, sheet.max_column + 1)]

        # Get sample rows
        for row_idx in range(start_row, min(start_row + num_rows, sheet.max_row + 1)):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_data[header] = cell.value
            sample_data.append(row_data)

        return sample_data

    def _get_cell_background_color(self, cell: Cell) -> str | None:
        """Extract background color from a cell."""
        try:
            fill = cell.fill
            if fill and isinstance(fill, PatternFill):
                if fill.fgColor and fill.fgColor.rgb:
                    color = str(fill.fgColor.rgb)
                    # Remove alpha channel if present (ARGB format)
                    if len(color) == 8:
                        color = color[2:]
                    return color.upper()
        except Exception:
            pass
        return None

    def _color_to_label(self, hex_color: str) -> str:
        """Convert hex color to semantic label."""
        if not hex_color:
            return "Unknown"

        hex_color = hex_color.upper().lstrip("#")

        # Direct match
        if hex_color in COLOR_LABELS:
            return COLOR_LABELS[hex_color]

        # Find closest color (simple RGB distance)
        try:
            r1 = int(hex_color[0:2], 16)
            g1 = int(hex_color[2:4], 16)
            b1 = int(hex_color[4:6], 16)

            min_distance = float("inf")
            closest_label = f"Custom({hex_color})"

            for known_hex, label in COLOR_LABELS.items():
                r2 = int(known_hex[0:2], 16)
                g2 = int(known_hex[2:4], 16)
                b2 = int(known_hex[4:6], 16)

                distance = ((r1 - r2) ** 2 + (g1 - g2) ** 2 + (b1 - b2) ** 2) ** 0.5

                if distance < min_distance and distance < 100:  # Threshold for "close enough"
                    min_distance = distance
                    closest_label = label

            return closest_label

        except (ValueError, IndexError):
            return f"Custom({hex_color})"


def manifest_to_dict(manifest: WorkbookManifest) -> dict:
    """Convert WorkbookManifest to JSON-serializable dictionary."""
    return {
        "file_path": manifest.file_path,
        "sheet_count": manifest.sheet_count,
        "sheet_names": manifest.sheet_names,
        "total_merged_regions": manifest.total_merged_regions,
        "total_sections": manifest.total_sections,
        "detected_colors": manifest.detected_colors,
        "sheets": {
            name: {
                "sheet_name": sheet.sheet_name,
                "total_rows": sheet.total_rows,
                "total_cols": sheet.total_cols,
                "merged_regions": [
                    {
                        "range": mr.range_string,
                        "value": mr.value,
                        "color": mr.color_label,
                        "row_span": mr.row_span,
                        "col_span": mr.col_span,
                    }
                    for mr in sheet.merged_regions
                ],
                "sections": [
                    {
                        "name": s.name,
                        "rows": f"{s.start_row}-{s.end_row}",
                        "cols": f"{s.start_col}-{s.end_col}",
                        "color": s.color_label,
                        "headers": s.column_headers,
                        "row_count": s.row_count,
                    }
                    for s in sheet.sections
                ],
                "color_regions": {color: len(cells) for color, cells in sheet.color_regions.items()},
                "empty_rows": sheet.empty_rows[:10],  # Limit for readability
                "header_row": sheet.header_row,
                "data_start_row": sheet.data_start_row,
                "sample_data": sheet.sample_data,
            }
            for name, sheet in manifest.sheets.items()
        },
    }
