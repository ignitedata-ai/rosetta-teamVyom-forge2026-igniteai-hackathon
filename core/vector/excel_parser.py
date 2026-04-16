"""Comprehensive Excel parser for extracting rich semantic information."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.logging import get_logger

logger = get_logger(__name__)


@dataclass
class CellInfo:
    """Information about a single cell."""

    address: str
    value: Any
    formula: Optional[str] = None
    data_type: str = "unknown"
    is_error: bool = False
    error_type: Optional[str] = None
    has_comment: bool = False
    comment: Optional[str] = None
    is_merged: bool = False
    style_info: dict = field(default_factory=dict)


@dataclass
class ColumnAnalysis:
    """Analysis of a single column."""

    name: str
    letter: str
    index: int
    data_type: str
    non_null_count: int
    null_count: int
    unique_count: int
    sample_values: list[Any]
    min_value: Optional[Any] = None
    max_value: Optional[Any] = None
    mean_value: Optional[float] = None
    has_formulas: bool = False
    formula_count: int = 0
    has_errors: bool = False
    error_count: int = 0
    inferred_purpose: Optional[str] = None

    def to_dict(self) -> dict:
        """Convert to serializable dictionary."""
        return {
            "name": self.name,
            "letter": self.letter,
            "index": self.index,
            "data_type": self.data_type,
            "non_null_count": self.non_null_count,
            "null_count": self.null_count,
            "unique_count": self.unique_count,
            "sample_values": [str(v)[:100] for v in self.sample_values[:5]],
            "min_value": str(self.min_value) if self.min_value is not None else None,
            "max_value": str(self.max_value) if self.max_value is not None else None,
            "mean_value": round(self.mean_value, 2) if self.mean_value is not None else None,
            "has_formulas": self.has_formulas,
            "formula_count": self.formula_count,
            "has_errors": self.has_errors,
            "error_count": self.error_count,
            "inferred_purpose": self.inferred_purpose,
        }


@dataclass
class SheetAnalysis:
    """Comprehensive analysis of a single sheet."""

    name: str
    row_count: int
    column_count: int
    data_range: str
    columns: list[ColumnAnalysis]
    formulas: list[dict]
    errors: list[dict]
    merged_cells: list[str]
    named_ranges: list[str]
    comments: list[dict]
    data_regions: list[dict]
    inferred_purpose: Optional[str] = None
    summary_statistics: dict = field(default_factory=dict)
    data_patterns: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        """Convert to serializable dictionary."""
        return {
            "name": self.name,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "data_range": self.data_range,
            "columns": [col.to_dict() for col in self.columns],
            "formula_count": len(self.formulas),
            "formulas": self.formulas[:20],  # Limit to first 20
            "error_count": len(self.errors),
            "errors": self.errors[:20],  # Limit to first 20
            "merged_cells_count": len(self.merged_cells),
            "merged_cells": self.merged_cells[:10],  # Limit
            "comments_count": len(self.comments),
            "comments": self.comments[:10],  # Limit
            "data_regions": self.data_regions,
            "inferred_purpose": self.inferred_purpose,
            "summary_statistics": self.summary_statistics,
            "data_patterns": self.data_patterns,
        }


@dataclass
class WorkbookAnalysis:
    """Complete analysis of an Excel workbook."""

    file_name: str
    sheet_count: int
    sheets: list[SheetAnalysis]
    total_formulas: int
    total_errors: int
    named_ranges: list[str]
    external_links: list[str]
    document_properties: dict
    relationships: list[dict]
    overall_purpose: Optional[str] = None

    def to_dict(self) -> dict:
        """Convert to serializable dictionary for database storage."""
        return {
            "file_name": self.file_name,
            "sheet_count": self.sheet_count,
            "sheets": [sheet.to_dict() for sheet in self.sheets],
            "total_formulas": self.total_formulas,
            "total_errors": self.total_errors,
            "named_ranges": self.named_ranges[:50],  # Limit
            "external_links": self.external_links[:20],  # Limit
            "document_properties": self.document_properties,
            "relationships": self.relationships[:50],  # Limit
            "overall_purpose": self.overall_purpose,
            # Summary for quick access
            "summary": {
                "total_rows": sum(s.row_count for s in self.sheets),
                "total_columns": sum(s.column_count for s in self.sheets),
                "has_formulas": self.total_formulas > 0,
                "has_errors": self.total_errors > 0,
                "formula_categories": self._get_formula_categories(),
                "error_types": self._get_error_types(),
                "column_purposes": self._get_column_purposes(),
            },
        }

    def _get_formula_categories(self) -> list[str]:
        """Get unique formula categories across all sheets."""
        categories = set()
        for sheet in self.sheets:
            for formula in sheet.formulas:
                if formula.get("category"):
                    categories.add(formula["category"])
        return list(categories)

    def _get_error_types(self) -> list[str]:
        """Get unique error types across all sheets."""
        error_types = set()
        for sheet in self.sheets:
            for error in sheet.errors:
                if error.get("error_type"):
                    error_types.add(error["error_type"])
        return list(error_types)

    def _get_column_purposes(self) -> dict[str, int]:
        """Get count of columns by inferred purpose."""
        purposes: dict[str, int] = {}
        for sheet in self.sheets:
            for col in sheet.columns:
                purpose = col.inferred_purpose or "general"
                purposes[purpose] = purposes.get(purpose, 0) + 1
        return purposes


class ExcelParser:
    """Comprehensive Excel parser that extracts rich semantic information."""

    # Common Excel error values
    ERROR_VALUES = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#GETTING_DATA"}

    # Common formula functions for categorization
    FORMULA_CATEGORIES = {
        "mathematical": ["SUM", "AVERAGE", "COUNT", "MIN", "MAX", "ROUND", "ABS", "SQRT"],
        "logical": ["IF", "AND", "OR", "NOT", "IFERROR", "IFS", "SWITCH"],
        "lookup": ["VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "XLOOKUP", "LOOKUP"],
        "text": ["CONCATENATE", "LEFT", "RIGHT", "MID", "LEN", "TRIM", "UPPER", "LOWER"],
        "date": ["DATE", "TODAY", "NOW", "YEAR", "MONTH", "DAY", "DATEDIF", "EDATE"],
        "financial": ["PMT", "FV", "PV", "NPV", "IRR", "RATE"],
        "statistical": ["STDEV", "VAR", "MEDIAN", "MODE", "PERCENTILE", "CORREL"],
    }

    def __init__(self):
        """Initialize the Excel parser."""
        pass

    def parse_workbook(
        self,
        file_content: bytes,
        file_name: str = "workbook.xlsx",
    ) -> WorkbookAnalysis:
        """Parse an Excel workbook and extract comprehensive information.

        Args:
            file_content: Raw file content
            file_name: Original file name

        Returns:
            Complete workbook analysis

        """
        try:
            # Load workbook with openpyxl for formula/style access
            wb = load_workbook(BytesIO(file_content), data_only=False)
            wb_data = load_workbook(BytesIO(file_content), data_only=True)

            sheets_analysis = []
            total_formulas = 0
            total_errors = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_data = wb_data[sheet_name]

                sheet_analysis = self._analyze_sheet(ws, ws_data, sheet_name)
                sheets_analysis.append(sheet_analysis)

                total_formulas += len(sheet_analysis.formulas)
                total_errors += len(sheet_analysis.errors)

            # Extract workbook-level information
            # In newer openpyxl, defined_names is a DefinedNameDict (dict-like)
            named_range_names = list(wb.defined_names.keys()) if wb.defined_names else []

            # Document properties
            doc_props = {}
            if wb.properties:
                doc_props = {
                    "title": wb.properties.title,
                    "subject": wb.properties.subject,
                    "creator": wb.properties.creator,
                    "created": str(wb.properties.created) if wb.properties.created else None,
                    "modified": str(wb.properties.modified) if wb.properties.modified else None,
                    "description": wb.properties.description,
                    "keywords": wb.properties.keywords,
                    "category": wb.properties.category,
                }

            # Infer overall purpose
            overall_purpose = self._infer_workbook_purpose(sheets_analysis, doc_props)

            # Detect relationships between sheets
            relationships = self._detect_sheet_relationships(sheets_analysis)

            workbook_analysis = WorkbookAnalysis(
                file_name=file_name,
                sheet_count=len(wb.sheetnames),
                sheets=sheets_analysis,
                total_formulas=total_formulas,
                total_errors=total_errors,
                named_ranges=named_range_names,
                external_links=[],  # Would need additional parsing
                document_properties=doc_props,
                relationships=relationships,
                overall_purpose=overall_purpose,
            )

            logger.info(
                "Workbook parsed successfully",
                file_name=file_name,
                sheet_count=len(sheets_analysis),
                total_formulas=total_formulas,
                total_errors=total_errors,
            )

            return workbook_analysis

        except Exception as e:
            logger.error(f"Error parsing workbook: {e}", exc_info=True)
            raise

    def _analyze_sheet(
        self,
        ws: Worksheet,
        ws_data: Worksheet,
        sheet_name: str,
    ) -> SheetAnalysis:
        """Analyze a single worksheet."""
        # Get dimensions
        min_row = ws.min_row or 1
        max_row = ws.max_row or 1
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1

        row_count = max_row - min_row + 1
        column_count = max_col - min_col + 1
        data_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

        # Analyze columns
        columns = self._analyze_columns(ws, ws_data, min_row, max_row, min_col, max_col)

        # Extract formulas
        formulas = self._extract_formulas(ws)

        # Extract errors
        errors = self._extract_errors(ws, ws_data)

        # Get merged cells
        merged_cells = [str(merged) for merged in ws.merged_cells.ranges]

        # Get comments
        comments = self._extract_comments(ws)

        # Detect data regions (headers, data areas, totals)
        data_regions = self._detect_data_regions(ws, ws_data, max_row, max_col)

        # Calculate summary statistics using pandas
        summary_statistics = self._calculate_summary_statistics(ws_data, min_row, max_row, min_col, max_col)

        # Detect data patterns
        data_patterns = self._detect_data_patterns(ws, ws_data, columns)

        # Infer sheet purpose
        inferred_purpose = self._infer_sheet_purpose(sheet_name, columns, formulas, data_regions, summary_statistics)

        return SheetAnalysis(
            name=sheet_name,
            row_count=row_count,
            column_count=column_count,
            data_range=data_range,
            columns=columns,
            formulas=formulas,
            errors=errors,
            merged_cells=merged_cells,
            named_ranges=[],  # Would need workbook-level access
            comments=comments,
            data_regions=data_regions,
            inferred_purpose=inferred_purpose,
            summary_statistics=summary_statistics,
            data_patterns=data_patterns,
        )

    def _analyze_columns(
        self,
        ws: Worksheet,
        ws_data: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> list[ColumnAnalysis]:
        """Analyze each column in the sheet."""
        columns = []

        for col_idx in range(min_col, max_col + 1):
            col_letter = get_column_letter(col_idx)

            # Get header (assume first row is header)
            header_cell = ws_data.cell(row=min_row, column=col_idx)
            header_name = str(header_cell.value) if header_cell.value else f"Column_{col_letter}"

            # Analyze column data
            values = []
            formula_count = 0
            error_count = 0

            for row_idx in range(min_row + 1, max_row + 1):  # Skip header
                cell = ws.cell(row=row_idx, column=col_idx)
                data_cell = ws_data.cell(row=row_idx, column=col_idx)

                # Check for formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

                # Check for error
                cell_value = data_cell.value
                if cell_value and str(cell_value) in self.ERROR_VALUES:
                    error_count += 1

                if cell_value is not None:
                    values.append(cell_value)

            # Determine data type
            data_type = self._infer_column_type(values)

            # Calculate statistics
            non_null_count = len(values)
            null_count = max_row - min_row - non_null_count
            unique_count = len(set(str(v) for v in values))

            # Sample values
            sample_values = values[:5] if values else []

            # Numeric statistics
            min_value = None
            max_value = None
            mean_value = None

            numeric_values = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if numeric_values:
                min_value = min(numeric_values)
                max_value = max(numeric_values)
                mean_value = sum(numeric_values) / len(numeric_values)

            # Infer purpose
            inferred_purpose = self._infer_column_purpose(header_name, data_type, sample_values)

            columns.append(
                ColumnAnalysis(
                    name=header_name,
                    letter=col_letter,
                    index=col_idx,
                    data_type=data_type,
                    non_null_count=non_null_count,
                    null_count=null_count,
                    unique_count=unique_count,
                    sample_values=sample_values,
                    min_value=min_value,
                    max_value=max_value,
                    mean_value=mean_value,
                    has_formulas=formula_count > 0,
                    formula_count=formula_count,
                    has_errors=error_count > 0,
                    error_count=error_count,
                    inferred_purpose=inferred_purpose,
                )
            )

        return columns

    def _infer_column_type(self, values: list) -> str:
        """Infer the data type of a column."""
        if not values:
            return "empty"

        # Filter out None values
        non_null_values = [v for v in values if v is not None]
        if not non_null_values:
            return "empty"

        type_counts = {
            "numeric": 0,
            "text": 0,
            "date": 0,
            "boolean": 0,
        }

        for val in non_null_values:
            if isinstance(val, bool):
                type_counts["boolean"] += 1
            elif isinstance(val, (int, float)):
                type_counts["numeric"] += 1
            elif hasattr(val, "strftime"):  # datetime-like
                type_counts["date"] += 1
            else:
                type_counts["text"] += 1

        # Return dominant type
        dominant_type = max(type_counts, key=type_counts.get)
        return dominant_type

    def _infer_column_purpose(
        self,
        header_name: str,
        data_type: str,
        sample_values: list,
    ) -> str:
        """Infer the purpose of a column based on its characteristics."""
        header_lower = header_name.lower()

        # Common column purpose patterns
        purpose_patterns = {
            "identifier": ["id", "code", "key", "number", "no", "num", "#"],
            "name": ["name", "title", "label", "description"],
            "date": ["date", "time", "timestamp", "created", "modified", "updated", "due"],
            "amount": ["amount", "total", "sum", "price", "cost", "value", "balance", "qty", "quantity"],
            "percentage": ["percent", "rate", "%", "ratio"],
            "status": ["status", "state", "flag", "active", "enabled"],
            "category": ["category", "type", "class", "group", "department"],
            "contact": ["email", "phone", "address", "city", "country", "zip"],
        }

        for purpose, patterns in purpose_patterns.items():
            for pattern in patterns:
                if pattern in header_lower:
                    return purpose

        # Infer from data type
        if data_type == "date":
            return "date"
        elif data_type == "numeric":
            return "numeric_value"
        elif data_type == "boolean":
            return "flag"

        return "general"

    def _extract_formulas(self, ws: Worksheet) -> list[dict]:
        """Extract all formulas from a worksheet."""
        formulas = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_info = {
                        "address": cell.coordinate,
                        "formula": cell.value,
                        "category": self._categorize_formula(cell.value),
                        "references": self._extract_cell_references(cell.value),
                    }
                    formulas.append(formula_info)

        return formulas

    def _categorize_formula(self, formula: str) -> str:
        """Categorize a formula based on its functions."""
        formula_upper = formula.upper()

        for category, functions in self.FORMULA_CATEGORIES.items():
            for func in functions:
                if func in formula_upper:
                    return category

        return "other"

    def _extract_cell_references(self, formula: str) -> list[str]:
        """Extract cell references from a formula."""
        # Pattern to match cell references like A1, $A$1, A1:B10, Sheet1!A1
        pattern = r"(?:[\w]+!)?\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?"
        matches = re.findall(pattern, formula, re.IGNORECASE)
        return matches

    def _extract_errors(self, ws: Worksheet, ws_data: Worksheet) -> list[dict]:
        """Extract all error cells from a worksheet."""
        errors = []

        for row_idx, row in enumerate(ws_data.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value and str(cell.value) in self.ERROR_VALUES:
                    # Get the original formula if any
                    original_cell = ws.cell(row=row_idx, column=col_idx)
                    formula = None
                    if original_cell.value and isinstance(original_cell.value, str) and original_cell.value.startswith("="):
                        formula = original_cell.value

                    errors.append(
                        {
                            "address": cell.coordinate,
                            "error_type": str(cell.value),
                            "formula": formula,
                            "description": self._get_error_description(str(cell.value)),
                        }
                    )

        return errors

    def _get_error_description(self, error_value: str) -> str:
        """Get a human-readable description of an Excel error."""
        descriptions = {
            "#DIV/0!": "Division by zero",
            "#N/A": "Value not available",
            "#NAME?": "Unrecognized formula name",
            "#NULL!": "Incorrect range reference",
            "#NUM!": "Invalid numeric value",
            "#REF!": "Invalid cell reference",
            "#VALUE!": "Wrong type of argument",
            "#GETTING_DATA": "Data is still loading",
        }
        return descriptions.get(error_value, "Unknown error")

    def _extract_comments(self, ws: Worksheet) -> list[dict]:
        """Extract all comments from a worksheet."""
        comments = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append(
                        {
                            "address": cell.coordinate,
                            "author": cell.comment.author,
                            "text": cell.comment.text,
                        }
                    )

        return comments

    def _detect_data_regions(
        self,
        ws: Worksheet,
        ws_data: Worksheet,
        max_row: int,
        max_col: int,
    ) -> list[dict]:
        """Detect different data regions in the sheet (headers, data, totals)."""
        regions = []

        if max_row < 1 or max_col < 1:
            return regions

        # Detect header row (usually row 1 with text values)
        header_row = 1
        header_values = []
        for col_idx in range(1, max_col + 1):
            cell = ws_data.cell(row=1, column=col_idx)
            if cell.value:
                header_values.append(str(cell.value))

        if header_values:
            regions.append(
                {
                    "type": "header",
                    "range": f"A1:{get_column_letter(max_col)}1",
                    "description": f"Header row with {len(header_values)} columns",
                    "values": header_values[:10],  # First 10 headers
                }
            )

        # Detect data region
        data_start_row = 2
        data_end_row = max_row

        # Check for totals row (last row with formulas or "Total" keyword)
        last_row_has_total = False
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=max_row, column=col_idx)
            cell_data = ws_data.cell(row=max_row, column=col_idx)

            if cell.value and isinstance(cell.value, str):
                if cell.value.startswith("=") and "SUM" in cell.value.upper():
                    last_row_has_total = True
                    break
            if cell_data.value and "total" in str(cell_data.value).lower():
                last_row_has_total = True
                break

        if last_row_has_total:
            data_end_row = max_row - 1
            regions.append(
                {
                    "type": "totals",
                    "range": f"A{max_row}:{get_column_letter(max_col)}{max_row}",
                    "description": "Totals/Summary row",
                }
            )

        if data_end_row >= data_start_row:
            regions.append(
                {
                    "type": "data",
                    "range": f"A{data_start_row}:{get_column_letter(max_col)}{data_end_row}",
                    "description": f"Data region with {data_end_row - data_start_row + 1} rows",
                    "row_count": data_end_row - data_start_row + 1,
                }
            )

        return regions

    def _calculate_summary_statistics(
        self,
        ws_data: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> dict:
        """Calculate summary statistics for numeric columns."""
        try:
            # Convert to pandas for easier analysis
            data = []
            headers = []

            # Get headers
            for col_idx in range(min_col, max_col + 1):
                cell = ws_data.cell(row=min_row, column=col_idx)
                headers.append(str(cell.value) if cell.value else f"Col_{col_idx}")

            # Get data
            for row_idx in range(min_row + 1, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                data.append(row_data)

            if not data:
                return {}

            df = pd.DataFrame(data, columns=headers)
            numeric_df = df.select_dtypes(include=["number"])

            if numeric_df.empty:
                return {"message": "No numeric columns found"}

            stats = {}
            for col in numeric_df.columns:
                col_stats = numeric_df[col].describe().to_dict()
                stats[col] = {
                    "count": int(col_stats.get("count", 0)),
                    "mean": round(col_stats.get("mean", 0), 2),
                    "std": round(col_stats.get("std", 0), 2),
                    "min": col_stats.get("min"),
                    "max": col_stats.get("max"),
                    "sum": round(numeric_df[col].sum(), 2),
                }

            return stats

        except Exception as e:
            logger.warning(f"Error calculating summary statistics: {e}")
            return {}

    def _detect_data_patterns(
        self,
        ws: Worksheet,
        ws_data: Worksheet,
        columns: list[ColumnAnalysis],
    ) -> list[str]:
        """Detect common data patterns in the sheet."""
        patterns = []

        # Check for time series data
        date_columns = [c for c in columns if c.data_type == "date"]
        if date_columns:
            patterns.append(f"Contains time-series data with {len(date_columns)} date column(s)")

        # Check for financial data
        financial_keywords = ["amount", "price", "cost", "total", "balance", "revenue", "expense"]
        financial_cols = [c for c in columns if any(kw in c.name.lower() for kw in financial_keywords)]
        if financial_cols:
            patterns.append(f"Contains financial data with columns: {', '.join(c.name for c in financial_cols[:5])}")

        # Check for calculated columns
        formula_columns = [c for c in columns if c.has_formulas]
        if formula_columns:
            patterns.append(f"{len(formula_columns)} column(s) contain calculated values (formulas)")

        # Check for high cardinality (likely IDs)
        id_columns = [c for c in columns if c.unique_count == c.non_null_count and c.non_null_count > 1]
        if id_columns:
            patterns.append(f"Unique identifier column(s) detected: {', '.join(c.name for c in id_columns[:3])}")

        # Check for categorical data
        categorical_cols = [c for c in columns if c.data_type == "text" and c.unique_count < 20 and c.non_null_count > 10]
        if categorical_cols:
            patterns.append(f"Categorical data detected in: {', '.join(c.name for c in categorical_cols[:3])}")

        return patterns

    def _infer_sheet_purpose(
        self,
        sheet_name: str,
        columns: list[ColumnAnalysis],
        formulas: list[dict],
        data_regions: list[dict],
        summary_statistics: dict,
    ) -> str:
        """Infer the purpose of a sheet based on its characteristics."""
        purposes = []

        # Check sheet name for clues
        name_lower = sheet_name.lower()
        if any(kw in name_lower for kw in ["summary", "overview", "dashboard"]):
            purposes.append("summary/dashboard")
        elif any(kw in name_lower for kw in ["data", "raw", "source"]):
            purposes.append("raw data storage")
        elif any(kw in name_lower for kw in ["calc", "analysis", "report"]):
            purposes.append("analysis/reporting")
        elif any(kw in name_lower for kw in ["config", "settings", "param"]):
            purposes.append("configuration")
        elif any(kw in name_lower for kw in ["lookup", "reference", "master"]):
            purposes.append("reference/lookup data")

        # Check for formula density
        data_region = next((r for r in data_regions if r["type"] == "data"), None)
        if data_region and formulas:
            row_count = data_region.get("row_count", 1)
            formula_density = len(formulas) / (row_count * len(columns)) if columns else 0
            if formula_density > 0.5:
                purposes.append("calculation sheet")
            elif formula_density > 0.1:
                purposes.append("mixed data and calculations")

        # Check column types
        numeric_cols = [c for c in columns if c.data_type == "numeric"]
        date_cols = [c for c in columns if c.data_type == "date"]

        if len(numeric_cols) > len(columns) * 0.7:
            purposes.append("numeric/financial data")
        if date_cols and numeric_cols:
            purposes.append("time-series tracking")

        if not purposes:
            purposes.append("general data storage")

        return ", ".join(purposes)

    def _infer_workbook_purpose(
        self,
        sheets: list[SheetAnalysis],
        doc_props: dict,
    ) -> str:
        """Infer the overall purpose of the workbook."""
        purposes = []

        # Check document properties
        if doc_props.get("title"):
            purposes.append(f"Title: {doc_props['title']}")
        if doc_props.get("subject"):
            purposes.append(f"Subject: {doc_props['subject']}")

        # Analyze sheet purposes
        sheet_purposes = [s.inferred_purpose for s in sheets if s.inferred_purpose]
        if sheet_purposes:
            purposes.append(f"Contains: {', '.join(set(sheet_purposes))}")

        # Count totals
        total_rows = sum(s.row_count for s in sheets)
        total_formulas = sum(len(s.formulas) for s in sheets)

        purposes.append(f"Multi-sheet workbook with {len(sheets)} sheets, {total_rows} total rows")

        if total_formulas > 100:
            purposes.append("Heavy calculation workbook")
        elif total_formulas > 10:
            purposes.append("Contains calculations")

        return "; ".join(purposes)

    def _detect_sheet_relationships(self, sheets: list[SheetAnalysis]) -> list[dict]:
        """Detect relationships between sheets based on formulas and references."""
        relationships = []

        for sheet in sheets:
            for formula in sheet.formulas:
                # Check if formula references other sheets
                formula_text = formula.get("formula", "")
                for other_sheet in sheets:
                    if other_sheet.name != sheet.name and other_sheet.name in formula_text:
                        relationships.append(
                            {
                                "from_sheet": sheet.name,
                                "to_sheet": other_sheet.name,
                                "type": "formula_reference",
                                "formula_location": formula.get("address"),
                            }
                        )

        return relationships
