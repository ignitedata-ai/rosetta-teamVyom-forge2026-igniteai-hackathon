"""Workbook parser — extracts full computational structure.

Uses openpyxl in two passes:
  1. data_only=False to access formulas
  2. data_only=True to access cached values
"""

from __future__ import annotations

import logging
import re
import uuid
from pathlib import Path
from typing import Any

import openpyxl

from .formula_parser import expand_refs, parse_formula
from .models import (
    CellModel,
    DependencyGraphSummary,
    NamedRangeModel,
    RegionModel,
    SheetModel,
    WorkbookModel,
)

log = logging.getLogger("rosetta.parser")


def _canon(sheet: str, coord: str) -> str:
    return f"{sheet}!{coord.replace('$', '')}"


_CURRENCY_SYMBOLS = ("$", "€", "£", "¥", "₹", "[$-")

# Label keywords that strongly imply a percent / rate semantic meaning.
# Used as a fallback when the cell's number_format is "General" (common in
# programmatically-generated workbooks where the author never set an explicit
# percent format).
_PERCENT_LABEL_HINTS = ("rate", "percent", "ratio", "margin", "%")

# Label keywords that strongly imply a currency semantic meaning. Used only
# as a fallback when number_format provides no signal and the value magnitude
# is consistent with money (>= 1).
_CURRENCY_LABEL_HINTS = (
    "revenue",
    "cost",
    "price",
    "gross",
    "profit",
    "income",
    "expense",
    "amount",
    "total",
    "budget",
    "addback",
    "payroll",
    "fee",
    "spend",
    "pvr",
    "allowance",
    "compensation",
)


def _infer_data_type(
    value: Any,
    number_format: str | None = None,
    *,
    label: str | None = None,
) -> str:
    """Infer a semantic data type from the cell's cached value, its Excel
    number format string, and optionally its semantic label.

    Priority:
      1. None → empty; bool → bool; datetime-like → date; string → string/error
      2. Explicit number_format (`%`, currency symbol, date tokens) wins
      3. Label-based fallback for `percent` / `currency` when format is General
         (common in programmatically-built fixtures)
      4. Default → number
    """
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if hasattr(value, "isoformat"):
        return "date"
    if isinstance(value, (int, float)):
        fmt = (number_format or "").strip()
        # Explicit format wins
        if fmt and fmt.lower() != "general":
            if "%" in fmt:
                return "percent"
            if any(sym in fmt for sym in _CURRENCY_SYMBOLS):
                return "currency"
            if _looks_like_date_format(fmt):
                return "date"
        # Label-based fallbacks
        if label:
            low = label.lower()
            # Percent: small fractional value AND rate-like label
            if abs(value) <= 1 and value != 0 and any(kw in low for kw in _PERCENT_LABEL_HINTS):
                return "percent"
            # Currency: numeric magnitude >= 1 AND money-like label
            if abs(value) >= 1 and any(kw in low for kw in _CURRENCY_LABEL_HINTS):
                return "currency"
        return "number"
    if isinstance(value, str):
        if value.startswith("#"):
            return "error"
        return "string"
    return "other"


def _looks_like_date_format(fmt: str) -> bool:
    """Return True if the number_format contains date/time tokens (y / m / d / h / s)
    *outside* quoted sections. Excel uses the same 'm' for minutes and months;
    we err on the side of calling anything with these tokens a date format,
    which is correct for the Dealer fixture.
    """
    in_quotes = False
    for ch in fmt:
        if ch == '"':
            in_quotes = not in_quotes
            continue
        if in_quotes:
            continue
        if ch in "ymdhs":
            return True
    return False


def _extract_named_ranges(wb_formulas) -> list[NamedRangeModel]:
    result: list[NamedRangeModel] = []
    # Workbook-level
    for dn in wb_formulas.defined_names.values() if isinstance(wb_formulas.defined_names, dict) else wb_formulas.defined_names:
        # openpyxl 3.1+: defined_names is a DefinedNameDict mapping name -> DefinedName
        if isinstance(dn, str):
            name_obj = wb_formulas.defined_names[dn]
            name = dn
        else:
            name_obj = dn
            name = name_obj.name
        raw = name_obj.value or name_obj.attr_text or ""
        resolved = _resolve_defined_name(raw)
        is_dynamic = bool(re.search(r"\b(OFFSET|INDIRECT)\s*\(", raw, re.IGNORECASE))
        result.append(
            NamedRangeModel(
                name=name,
                scope="workbook",
                raw_value=raw,
                resolved_refs=resolved,
                is_dynamic=is_dynamic,
            )
        )
    # Sheet-scoped
    for sheet in wb_formulas.worksheets:
        for name, name_obj in sheet.defined_names.items() if hasattr(sheet.defined_names, "items") else []:
            raw = name_obj.value or ""
            resolved = _resolve_defined_name(raw)
            is_dynamic = bool(re.search(r"\b(OFFSET|INDIRECT)\s*\(", raw, re.IGNORECASE))
            result.append(
                NamedRangeModel(
                    name=name,
                    scope=sheet.title,
                    raw_value=raw,
                    resolved_refs=resolved,
                    is_dynamic=is_dynamic,
                )
            )
    return result


def _resolve_defined_name(raw: str) -> list[str]:
    """Parse a defined name target like Sheet!$A$1 or Sheet!$A$1:$B$10 into canonical refs."""
    if not raw:
        return []
    # Strip leading = if present
    v = raw.strip()
    if v.startswith("="):
        v = v[1:]
    # Handle quoted and unquoted sheet prefixes
    m = re.match(r"(?:'([^']+)'|([A-Za-z_][\w\.]*))!(.+)$", v)
    if not m:
        return []
    sheet = m.group(1) or m.group(2)
    rest = m.group(3).strip()
    rest_clean = rest.replace("$", "")
    return [f"{sheet}!{rest_clean}"]


def _detect_regions(sheet, max_row: int, max_col: int) -> list[RegionModel]:
    """Heuristic region detector: header / data / blank / subtotal / calculation."""
    regions: list[RegionModel] = []
    if max_row == 0:
        return regions

    row_kinds: list[str] = []
    for r in range(1, max_row + 1):
        non_empty = 0
        has_formula = False
        has_text_only = False
        total_numeric = 0
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            non_empty += 1
            if isinstance(v, str) and v.startswith("="):
                has_formula = True
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                total_numeric += 1
            elif isinstance(v, str):
                has_text_only = True
        if non_empty == 0:
            row_kinds.append("blank")
        elif has_formula and total_numeric == 0 and non_empty <= 2:
            row_kinds.append("calculation")
        elif has_formula:
            row_kinds.append("data")
        elif total_numeric > 0 and has_text_only:
            row_kinds.append("data")
        elif has_text_only and total_numeric == 0:
            row_kinds.append("header")
        else:
            row_kinds.append("data")

    # Coalesce
    i = 0
    while i < len(row_kinds):
        kind = row_kinds[i]
        j = i
        while j + 1 < len(row_kinds) and row_kinds[j + 1] == kind:
            j += 1
        regions.append(
            RegionModel(
                type=kind if kind in ("header", "data", "blank", "calculation") else "data",
                rows=(i + 1, j + 1),
            )
        )
        i = j + 1
    return regions


def _nearest_row_label(sheet, row: int) -> str | None:
    """Look at column A for a label on this row (common layout)."""
    v = sheet.cell(row=row, column=1).value
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None


def _column_header(sheet, col: int, max_header_row: int = 3) -> str | None:
    """Return header text for a column by scanning the first 1-3 rows."""
    for r in range(1, max_header_row + 1):
        v = sheet.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _semantic_label(sheet, row: int, col: int) -> str | None:
    """Combine column header and row label to produce a rich semantic label.

    Priority: row label in col A (if present and this cell isn't in col A)
              combined with column header → "<RowLabel> — <ColHeader>"
              else column header alone, else row label alone.
    """
    row_lbl = _nearest_row_label(sheet, row) if col > 1 else None
    col_hdr = _column_header(sheet, col) if row > 1 else None
    if row_lbl and col_hdr and row_lbl != col_hdr:
        return f"{row_lbl} — {col_hdr}"
    return row_lbl or col_hdr


def parse_workbook(path: str | Path, workbook_id: str | None = None) -> WorkbookModel:
    path = Path(path)
    wid = workbook_id or f"wb_{uuid.uuid4().hex[:10]}"
    log.info("Parsing workbook %s", path)

    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)

    named_ranges = _extract_named_ranges(wb_f)
    nr_names = [nr.name for nr in named_ranges]

    sheets: list[SheetModel] = []
    cells: dict[str, CellModel] = {}

    for ws_f in wb_f.worksheets:
        name = ws_f.title
        ws_v = wb_v[name]
        hidden = ws_f.sheet_state in ("hidden", "veryHidden")
        max_row = ws_f.max_row or 0
        max_col = ws_f.max_column or 0

        # Merged cells
        merged = [str(mr) for mr in ws_f.merged_cells.ranges]

        # Hidden rows/cols
        hidden_rows = [r for r, dim in ws_f.row_dimensions.items() if getattr(dim, "hidden", False)]
        hidden_cols = [c for c, dim in ws_f.column_dimensions.items() if getattr(dim, "hidden", False)]

        regions = _detect_regions(ws_f, max_row, max_col) if max_row > 0 else []

        sheet_model = SheetModel(
            name=name,
            hidden=hidden,
            max_row=max_row,
            max_col=max_col,
            merged_cells=merged,
            hidden_rows=hidden_rows,
            hidden_cols=hidden_cols,
            regions=regions,
            formula_count=0,
        )

        # Iterate cells
        if max_row > 0 and max_col > 0:
            for row in ws_f.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    fval = cell.value
                    vval = ws_v.cell(row=cell.row, column=cell.column).value
                    if fval is None and vval is None:
                        continue
                    coord = cell.coordinate
                    ref = _canon(name, coord)
                    formula = None
                    is_formula = isinstance(fval, str) and fval.startswith("=")
                    if is_formula:
                        formula = fval[1:]
                        sheet_model.formula_count += 1

                    label = _semantic_label(ws_f, cell.row, cell.column)

                    # Read number_format from the formula-pass workbook (more reliable
                    # than the data_only one for format strings).
                    number_format = getattr(cell, "number_format", None)
                    cm = CellModel(
                        sheet=name,
                        coord=coord,
                        ref=ref,
                        value=vval,
                        formula=formula,
                        data_type=_infer_data_type(vval, number_format, label=label),
                        is_hardcoded=not is_formula and vval is not None,
                        semantic_label=label,
                    )
                    if formula:
                        pf = parse_formula(formula, name, nr_names)
                        # Expand refs to individual cells / logical ranges
                        deps = expand_refs(pf.refs)
                        # Include named-range-resolved refs
                        for nr_used in pf.named_ranges:
                            nr_obj = next((n for n in named_ranges if n.name.upper() == nr_used.upper()), None)
                            if nr_obj:
                                for rr in nr_obj.resolved_refs:
                                    if ":" in rr:
                                        sheet_, rng_ = rr.split("!", 1)
                                        start, end = rng_.split(":", 1)
                                        from .formula_parser import expand_range

                                        deps.extend(expand_range(sheet_, start, end))
                                    else:
                                        deps.append(rr)
                        cm.depends_on = list(dict.fromkeys(deps))
                        cm.named_ranges_used = pf.named_ranges
                        cm.formula_type = pf.formula_type
                        cm.is_volatile = pf.is_volatile
                        cm.is_hardcoded = False
                    cells[ref] = cm
                    sheet_model.cell_refs.append(ref)

        # Resolve named range current values
        sheets.append(sheet_model)

    # Populate named range current values
    for nr in named_ranges:
        if len(nr.resolved_refs) == 1 and ":" not in nr.resolved_refs[0]:
            c = cells.get(nr.resolved_refs[0])
            if c:
                nr.current_value = c.value

    # Build reverse deps
    for ref, cm in cells.items():
        for d in cm.depends_on:
            if d in cells:
                if ref not in cells[d].depended_by:
                    cells[d].depended_by.append(ref)

    # Graph summary
    total_formula = sum(1 for c in cells.values() if c.formula)
    cross_sheet_edges = 0
    for cm in cells.values():
        for d in cm.depends_on:
            if "!" in d and d.split("!", 1)[0] != cm.sheet:
                cross_sheet_edges += 1

    # Detect circular refs via simple DFS, then enrich each with any
    # author comment from cells in the chain (authoritative for whether
    # a cycle is intentional vs accidental).
    circular = _detect_circular(cells)
    _enrich_circular_with_comments(wb_f, circular)

    # Max depth (bounded)
    max_depth = _approx_max_depth(cells)

    graph_summary = DependencyGraphSummary(
        total_formula_cells=total_formula,
        max_depth=max_depth,
        cross_sheet_edges=cross_sheet_edges,
        circular_references=circular,
    )

    # Parse pivot tables from the raw .xlsx XML and attach to their host sheets.
    # Failures here are non-fatal — pivots are enrichment, not critical path.
    try:
        from .pivot_parser import parse_pivot_tables

        pivots_by_sheet = parse_pivot_tables(path)
        sheets_by_name = {s.name: s for s in sheets}
        for sheet_name, pivots in pivots_by_sheet.items():
            target = sheets_by_name.get(sheet_name)
            if target is not None:
                target.pivot_tables = pivots
    except Exception as e:
        log.warning("pivot-table parsing failed for %s: %s", path, e)

    wb = WorkbookModel(
        workbook_id=wid,
        filename=path.name,
        sheets=sheets,
        named_ranges=named_ranges,
        cells=cells,
        graph_summary=graph_summary,
    )
    # If formulas were not pre-computed by Excel (common when files are
    # programmatically generated), run our evaluator to fill in cached values
    # so downstream queries / what-if / traces have numbers to show.
    _precompute_missing_values(wb)
    # Re-populate named range current values after compute
    for nr in wb.named_ranges:
        if len(nr.resolved_refs) == 1 and ":" not in nr.resolved_refs[0]:
            c = wb.cells.get(nr.resolved_refs[0])
            if c:
                nr.current_value = c.value
    log.info(
        "Parsed %s: %d sheets, %d cells, %d formulas, %d named ranges",
        path.name,
        len(sheets),
        len(cells),
        total_formula,
        len(named_ranges),
    )
    return wb


def _precompute_missing_values(wb: WorkbookModel) -> None:
    """If formula cells have no cached value (e.g. newly-written by openpyxl),
    evaluate them using Rosetta's evaluator so that values flow through.
    """
    from .evaluator import Evaluator

    # Only engage if a noticeable fraction of formula cells have None values
    formula_cells = [c for c in wb.cells.values() if c.formula]
    if not formula_cells:
        return
    missing = [c for c in formula_cells if c.value is None]
    if len(missing) / max(len(formula_cells), 1) < 0.3:
        return  # Excel-calculated file — trust cached values
    ev = Evaluator(wb)
    for c in formula_cells:
        if c.value is None:
            v = ev.value_of(c.ref)
            if v is not None:
                c.value = v
                # Re-infer using whatever format was previously stored on the
                # CellModel; we don't have the openpyxl cell here, so pass
                # through the existing data_type as a hint via the number_format
                # sentinel (None → plain value-based inference).
                c.data_type = _infer_data_type(v, None) if c.data_type in (None, "empty") else c.data_type


def _detect_circular(cells: dict[str, CellModel]):
    from .models import CircularRef

    WHITE, GRAY, BLACK = 0, 1, 2
    color = dict.fromkeys(cells, WHITE)
    stack: list[str] = []
    cycles: list[CircularRef] = []

    def dfs(node: str):
        if node not in cells:
            return
        if color[node] == GRAY:
            # cycle found
            if node in stack:
                idx = stack.index(node)
                chain = stack[idx:] + [node]
                if not any(set(chain[:-1]) == set(c.chain[:-1]) for c in cycles):
                    # Intentional defaults to False here; the comment-enrichment
                    # pass in parse_workbook promotes it to True (with an
                    # author_comment) when the workbook author left a note on
                    # a cell in the cycle.
                    cycles.append(
                        CircularRef(
                            chain=chain,
                            intentional=False,
                            note="Circular dependency detected. No author note found — may be intentional (iterative calc) or a bug.",
                        )
                    )
            return
        if color[node] == BLACK:
            return
        color[node] = GRAY
        stack.append(node)
        cell = cells[node]
        for d in cell.depends_on:
            if "!" in d and ":" not in d.split("!", 1)[1]:
                dfs(d)
        stack.pop()
        color[node] = BLACK

    for k in cells:
        if color[k] == WHITE:
            dfs(k)
    return cycles


# Keywords in an author's cell comment that promote a cycle to `intentional`.
_INTENTIONAL_COMMENT_KEYWORDS = (
    "circular",
    "iterative",
    "intentional",
    "by design",
    "expected",
)


def _enrich_circular_with_comments(wb_f, cycles) -> None:
    """For each detected cycle, look for evidence that the cycle is intentional:

      Primary:  an openpyxl cell comment on any cell in the chain containing
                an intent keyword (`circular`, `iterative`, `intentional`,
                `by design`, `expected`). This is authoritative — comments
                are explicit author annotations.

      Fallback: Excel's workbook-level iterative-calculation setting
                (`wb.calculation.iterate == True`). When set, Excel is
                configured to allow circular references and converge — a
                strong signal that at least the top-level cycles are
                intentional. We promote but mark it as inferred, not
                authored.
    """
    if not cycles:
        return
    # Primary pass: look for authored comments on cells in each cycle.
    for cr in cycles:
        for cell_ref in cr.chain:
            if "!" not in cell_ref:
                continue
            sheet_name, coord = cell_ref.split("!", 1)
            try:
                ws = wb_f[sheet_name]
            except KeyError:
                continue
            try:
                cell = ws[coord]
            except Exception:
                continue
            cmt = getattr(cell, "comment", None)
            if cmt is None:
                continue
            text = (cmt.text or "").strip()
            if not text:
                continue
            low = text.lower()
            if any(kw in low for kw in _INTENTIONAL_COMMENT_KEYWORDS):
                cr.intentional = True
                cr.author_comment = text
                cr.commented_ref = cell_ref
                cr.comment_author = getattr(cmt, "author", None)
                cr.note = f"Marked intentional by the author's comment on {cell_ref}."
                break  # one confirming comment is enough

    # Fallback: iterative-calc workbook setting. If enabled, Excel is actively
    # permitting circular refs to converge. Promote any still-ambiguous cycle.
    iterative_enabled = False
    try:
        iterative_enabled = bool(getattr(wb_f.calculation, "iterate", False))
    except Exception:
        iterative_enabled = False
    if iterative_enabled:
        for cr in cycles:
            if cr.intentional:
                continue
            cr.intentional = True
            cr.note = (
                "Iterative calculation is enabled on this workbook, so Excel is "
                "configured to resolve circular references by repeated evaluation. "
                "The cycle is treated as intentional (no author comment found)."
            )


def _approx_max_depth(cells: dict[str, CellModel]) -> int:
    """Memoized DFS for max dependency depth. Cycles clipped."""
    memo: dict[str, int] = {}

    def d(ref: str, visiting: set[str]) -> int:
        if ref in memo:
            return memo[ref]
        if ref in visiting:
            return 0
        c = cells.get(ref)
        if not c or not c.formula:
            memo[ref] = 0
            return 0
        visiting.add(ref)
        best = 0
        for dep in c.depends_on:
            if ":" in dep.split("!", 1)[-1]:
                continue
            best = max(best, 1 + d(dep, visiting))
        visiting.discard(ref)
        memo[ref] = best
        return best

    mx = 0
    for k in cells:
        mx = max(mx, d(k, set()))
    return mx
